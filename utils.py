import json
import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Alignment, PatternFill
import os

def createjsonobject(content):
 
# Format each block as a complete JSON object
    formatted_objects = []
    for i, obj in enumerate(content):
        # Add braces around each split object to make it a valid JSON object
        if not obj.startswith('{'):
            obj = '{' + obj
        if not obj.endswith('}'):
            obj = obj + '}'
        
        # Parse the object to confirm it's valid JSON
        try:
            parsed_obj = json.loads(obj)
            formatted_objects.append(parsed_obj)
        except json.JSONDecodeError as e:
            print(f"Error parsing object {i}: {e}")
    return formatted_objects            

def txttojson(input_txt):
    # Load the file and split by line to handle each object individually
    with open(input_txt, "r") as file:
        lines = file.readlines()
    
    json_objects = []
    current_object = ""
    
    for line in lines:
        current_object += line
        # Detect end of an object by checking if line ends with '}' and the object seems valid
        if line.strip() == "}," or line.strip() == "}":
            # Try parsing the accumulated JSON object
            try:
                json_obj = json.loads(current_object)
                json_objects.append(json_obj)
            except json.JSONDecodeError as e:
                print(f"Skipping invalid JSON object due to error: {e}")
            
            # Reset the current object accumulator
            current_object = ""
    
    return json_objects


def formatting_excel(unformatedexcel,blob_name):
    workbook = openpyxl.load_workbook(unformatedexcel)
    sheet = workbook.active

    # Remove '.pdf' extension from blob_name and add '.xlsx' extension
    base_blob_name = blob_name.replace(".pdf", "")  # Remove .pdf from blob name
    output_path = os.path.join("Insights", f"Insights_{base_blob_name}.xlsx")  # Save to "Insights" folder

    # Set column widths
    column_widths = {
        "A": 30,  # Adjust the width for the first column
        "B": 60, 
        "C": 40,
        "D": 20,
        "E": 20,
        "F": 20,
        "G": 20,
        "H": 20,
        "I": 30,
        "J": 90,  # Assuming "J" is "clause_text"
        "K": 90  # Adjust width for other columns as needed
    }

    # Set widths for each specified column
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Define header height and background color
    header_row = 1  # Assuming the first row is the header
    sheet.row_dimensions[header_row].height = 30  # Set the desired header height

    # Define the fill for the background color
    header_fill = PatternFill(start_color="78A2CC", end_color="78A2CC", fill_type="solid")

    # Apply background color to each cell in the header row
    for cell in sheet[header_row]:
        cell.fill = header_fill

    # Set height for all other rows
    other_row_height = 140  # Adjust this to the desired height for all other rows
    for row in range(2, sheet.max_row + 1):  # Start from the second row to the last row
        sheet.row_dimensions[row].height = other_row_height

    # Define columns to exclude from center alignment
    exclude_columns = ["J", "K"]  # Assuming "J" is "clause_text", "L" is "notes", "M" is "task_description"

    # Apply alignment settings
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            # Check if cell column is in exclude list; if so, wrap text and top-align
            if cell.column_letter in exclude_columns:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                # Center-align both horizontally and vertically
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save the workbook with a new name to avoid overwriting the original
    os.remove(f"{base_blob_name}.txt")
    os.remove(unformatedexcel)
    workbook.save(output_path)

def json_to_excel(formatted_json, output_excel_file,blob_name):
    # Process each item in the data
    for item in formatted_json:
        # Check if 'Clause Text' is a list
        if isinstance(item.get('Clause Text'), list):
            # Remove any occurrences of "]" and join the remaining text with newlines
            clean_clauses = [clause for clause in item['Clause Text'] if clause != "]"]
            item['Clause Text'] = "\n".join(clean_clauses)
    
    # Create a DataFrame from the loaded JSON data
    df = pd.DataFrame(formatted_json)
    
    # Write the DataFrame to an Excel file
    df.to_excel(output_excel_file, index=False)
    formatting_excel(output_excel_file,blob_name)